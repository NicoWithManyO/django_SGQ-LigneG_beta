import os
from django.http import FileResponse, JsonResponse
from django.views.decorators.http import require_http_methods
from rest_framework.decorators import api_view
from .services import RollExcelExporter


@api_view(['GET'])
def download_rolls_export(request):
    """Télécharger le fichier Excel des rouleaux."""
    try:
        exporter = RollExcelExporter()
        filepath = exporter.get_export_path()
        
        if os.path.exists(filepath):
            response = FileResponse(
                open(filepath, 'rb'),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename="rolls_export.xlsx"'
            return response
        else:
            return JsonResponse({
                'error': 'Aucun fichier d\'export trouvé'
            }, status=404)
            
    except Exception as e:
        return JsonResponse({
            'error': str(e)
        }, status=500)


@api_view(['GET'])
def export_status(request):
    """Obtenir le statut de l'export (nombre de lignes, dernière mise à jour)."""
    try:
        exporter = RollExcelExporter()
        filepath = exporter.get_export_path()
        
        if os.path.exists(filepath):
            from openpyxl import load_workbook
            wb = load_workbook(filepath, read_only=True)
            ws = wb.active
            row_count = ws.max_row - 1  # -1 pour les en-têtes
            wb.close()
            
            stats = os.stat(filepath)
            last_modified = stats.st_mtime
            
            return JsonResponse({
                'exists': True,
                'row_count': row_count,
                'last_modified': last_modified,
                'file_size': stats.st_size,
                'filepath': filepath
            })
        else:
            return JsonResponse({
                'exists': False,
                'row_count': 0
            })
            
    except Exception as e:
        return JsonResponse({
            'error': str(e)
        }, status=500)