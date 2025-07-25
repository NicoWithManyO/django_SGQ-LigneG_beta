import os
import shutil
from datetime import datetime
from django.conf import settings
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from production.models import Roll, CurrentProfile


class RollExcelExporter:
    """Service pour exporter les rouleaux dans un fichier Excel."""
    
    def __init__(self):
        self.excel_dir = os.path.join(settings.MEDIA_ROOT, 'exports')
        self.filename = 'rolls_export.xlsx'
        self.filepath = os.path.join(self.excel_dir, self.filename)
        self.max_rows = 10000  # Rotation après 10000 lignes
        
        # Créer le répertoire si nécessaire
        os.makedirs(self.excel_dir, exist_ok=True)
        
        # En-têtes des colonnes - TOUS les champs du modèle Roll
        self.headers = [
            'ID', 'ID Rouleau', 'OF', 'Numéro', 'Date/Heure', 'Opérateur', 'Shift',
            'Longueur (m)', 'Statut', 'Destination', 'Masse Tube (g)', 'Masse Totale (g)', 
            'Masse Nette (g)', 'Défauts Bloquants', 'Problèmes Épaisseur',
            'Épaisseur Moy. Gauche', 'Épaisseur Moy. Droite', 'Grammage Calc.',
            'Défauts', 'Profil', 'Commentaire'
        ]
    
    def _create_workbook(self):
        """Créer un nouveau fichier Excel avec les en-têtes."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Rouleaux"
        
        # Style pour les en-têtes
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center")
        
        # Ajouter les en-têtes
        for col, header in enumerate(self.headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            
        # Largeur des colonnes
        column_widths = {
            'A': 8,   # ID
            'B': 20,  # ID Rouleau
            'C': 10,  # OF
            'D': 10,  # Numéro
            'E': 20,  # Date/Heure
            'F': 20,  # Opérateur
            'G': 25,  # Shift
            'H': 12,  # Longueur (m)
            'I': 15,  # Statut
            'J': 15,  # Destination
            'K': 12,  # Masse Tube
            'L': 12,  # Masse Totale
            'M': 12,  # Masse Nette
            'N': 15,  # Défauts Bloquants
            'O': 18,  # Problèmes Épaisseur
            'P': 18,  # Épaisseur Moy. Gauche
            'Q': 18,  # Épaisseur Moy. Droite
            'R': 15,  # Grammage Calc.
            'S': 30,  # Défauts
            'T': 20,  # Profil
            'U': 30,  # Commentaire
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
            
        return wb
    
    def _get_roll_data(self, roll):
        """Extraire TOUTES les données d'un rouleau pour l'export."""
        # Défauts
        defects = roll.defects.all()
        defects_str = ', '.join([
            f"{d.defect_type.name} à {d.meter_position}m" 
            for d in defects
        ])
        
        # Profil - récupérer depuis CurrentProfile ou depuis l'OF
        profile_name = ''
        
        # Essayer depuis CurrentProfile d'abord
        current_profile = CurrentProfile.objects.first()
        if current_profile and current_profile.profile:
            profile_name = current_profile.profile.name
        
        # Sinon essayer depuis l'OF
        if not profile_name and roll.fabrication_order and hasattr(roll.fabrication_order, 'profile'):
            profile_name = roll.fabrication_order.profile.name
        
        return [
            roll.id,  # ID unique de la base de données
            roll.roll_id,
            roll.fabrication_order.order_number if roll.fabrication_order else '',
            roll.roll_number or '',
            roll.created_at.strftime('%Y-%m-%d %H:%M:%S'),
            f"{roll.shift.operator.first_name} {roll.shift.operator.last_name}" if roll.shift and roll.shift.operator else '',
            roll.shift.shift_id if roll.shift else '',
            float(roll.length) if roll.length else 0,
            roll.status or 'CONFORME',
            roll.destination or 'PRODUCTION',
            float(roll.tube_mass) if roll.tube_mass else 0,
            float(roll.total_mass) if roll.total_mass else 0,
            float(roll.net_mass) if roll.net_mass else 0,
            'Oui' if roll.has_blocking_defects else 'Non',
            'Oui' if roll.has_thickness_issues else 'Non',
            float(roll.avg_thickness_left) if roll.avg_thickness_left else '',
            float(roll.avg_thickness_right) if roll.avg_thickness_right else '',
            float(roll.grammage_calc) if roll.grammage_calc else '',
            defects_str,
            profile_name,
            roll.comment or ''
        ]
    
    def _check_rotation(self):
        """Vérifier si le fichier doit être archivé."""
        if os.path.exists(self.filepath):
            wb = load_workbook(self.filepath)
            ws = wb.active
            row_count = ws.max_row
            wb.close()
            
            if row_count >= self.max_rows:
                # Archiver le fichier
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                archive_name = f'rolls_export_{timestamp}.xlsx'
                archive_path = os.path.join(self.excel_dir, 'archives', archive_name)
                
                # Créer le répertoire d'archives
                os.makedirs(os.path.dirname(archive_path), exist_ok=True)
                
                # Déplacer le fichier
                shutil.move(self.filepath, archive_path)
                return True
        return False
    
    def export_roll(self, roll, update=True):
        """Ajouter ou mettre à jour un rouleau dans le fichier Excel."""
        try:
            # Vérifier la rotation
            self._check_rotation()
            
            # Charger ou créer le fichier
            if os.path.exists(self.filepath):
                wb = load_workbook(self.filepath)
                ws = wb.active
            else:
                wb = self._create_workbook()
                ws = wb.active
            
            # Récupérer les données du rouleau
            row_data = self._get_roll_data(roll)
            
            # Chercher si le rouleau existe déjà (par ID en colonne A)
            roll_row = None
            if update:
                for row_num in range(2, ws.max_row + 1):  # Commencer à 2 pour skip les en-têtes
                    if ws.cell(row=row_num, column=1).value == roll.id:
                        roll_row = row_num
                        break
            
            if roll_row:
                # Mettre à jour la ligne existante
                for col, value in enumerate(row_data, 1):
                    ws.cell(row=roll_row, column=col, value=value)
            else:
                # Ajouter une nouvelle ligne
                ws.append(row_data)
                roll_row = ws.max_row
            
            # Style pour les rouleaux non conformes
            if roll.status != 'CONFORME':
                fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
                for col in range(1, len(self.headers) + 1):
                    ws.cell(row=roll_row, column=col).fill = fill
            else:
                # Retirer le style rouge si le rouleau est devenu conforme
                for col in range(1, len(self.headers) + 1):
                    ws.cell(row=roll_row, column=col).fill = PatternFill()
            
            # Sauvegarder
            wb.save(self.filepath)
            wb.close()
            
            return True, self.filepath
            
        except Exception as e:
            return False, str(e)
    
    def get_export_path(self):
        """Retourner le chemin du fichier Excel."""
        return self.filepath